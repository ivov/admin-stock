from datetime import datetime
import os
from collections import defaultdict
from PyQt5 import QtWidgets, QtGui, QtCore
from openpyxl import load_workbook
from utils import utils_collection as utils
from utils import db_manager
from utils.styling import generic_title_style, generic_groupbox_normal_style
from utils import tests
from widgets.line_item_close_button import LineItemCloseButton
from widgets.spec_fields import AutocapField, StockNumberField
from widgets.message_boxes import WarningBox


class ProductionReport(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(ProductionReport, self).__init__(parent)
        self.setWindowFlags(
            QtCore.Qt.Dialog
            | QtCore.Qt.CustomizeWindowHint
            | QtCore.Qt.WindowCloseButtonHint
        )
        self.setFixedWidth(410)

        self.entered_products_and_amounts = {}
        self.needed_comps_and_amounts = {}

        title = QtWidgets.QLabel("Informe de producción")
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet(generic_title_style)

        code_label = QtWidgets.QLabel("Código:")
        self.code_field = AutocapField("K12345...")

        code_regex = QtCore.QRegExp("^[KBTkbt]{1}[0-9]{1,7}")
        code_validator = QtGui.QRegExpValidator(code_regex, self.code_field)

        self.code_field.setValidator(code_validator)
        self.code_field.setFixedWidth(110)
        self.code_field.setFocus()

        code_section = QtWidgets.QHBoxLayout()
        code_section.addStretch()
        code_section.addWidget(code_label)
        code_section.addWidget(self.code_field)
        code_section.addStretch()

        groupbox = QtWidgets.QGroupBox("Productos a fabricar")
        groupbox.setStyleSheet(generic_groupbox_normal_style)
        groupbox.setMinimumWidth(252)

        self.add_button = QtWidgets.QPushButton("+ Agregar producto")
        self.add_button.setShortcut("Alt+a")
        self.products_holder_section = QtWidgets.QVBoxLayout()

        groupbox_inner_section = QtWidgets.QVBoxLayout()
        groupbox_inner_section_1 = QtWidgets.QHBoxLayout()
        groupbox_inner_section_1.addWidget(self.add_button)
        groupbox_inner_section_2 = QtWidgets.QHBoxLayout()
        groupbox_inner_section_2.addLayout(self.products_holder_section)
        groupbox_inner_section.addLayout(groupbox_inner_section_1)
        groupbox_inner_section.addLayout(groupbox_inner_section_2)
        groupbox.setLayout(groupbox_inner_section)
        groupbox_section = QtWidgets.QHBoxLayout()
        groupbox_section.addStretch()
        groupbox_section.addWidget(groupbox)
        groupbox_section.addStretch()

        self.execute_button = QtWidgets.QPushButton("Calcular »")
        self.execute_button.setShortcut("Alt+c")
        self.execute_button.setFixedWidth(150)

        self.back_button = QtWidgets.QPushButton("« Volver")
        self.back_button.setShortcut("Alt+v")
        self.back_button.setFixedWidth(150)

        self.bottom_section = QtWidgets.QHBoxLayout()
        self.bottom_section.addWidget(self.back_button)
        self.bottom_section.addWidget(self.execute_button)

        self.layout = QtWidgets.QVBoxLayout()
        self.layout.setSizeConstraint(QtWidgets.QLayout.SetFixedSize)
        self.layout.addWidget(title)
        self.layout.addLayout(code_section)
        self.layout.addLayout(groupbox_section)
        self.layout.addLayout(self.bottom_section)

        self.setLayout(self.layout)

        self.add_button.clicked.connect(self.add_line_item)
        self.back_button.clicked.connect(self.close)
        self.execute_button.clicked.connect(self.on_execute_clicked)

    def add_line_item(self):
        product_field = AutocapField("Producto...")
        product_field.setFixedWidth(120)
        product_field.set_completer(source="recipes")
        product_field.setFocus()

        value_field = StockNumberField("Cantidad...")
        value_field.setFixedWidth(80)

        close_button = LineItemCloseButton(holder=self.products_holder_section)

        single_line_section = QtWidgets.QHBoxLayout()
        single_line_section.addWidget(product_field)
        single_line_section.addWidget(value_field)
        single_line_section.addWidget(close_button)
        self.products_holder_section.addLayout(single_line_section)

    def on_execute_clicked(self):
        if self.execute_button.text() == "Calcular »":
            if self.test_if_there_are_no_products():
                return
            if self.test_if_there_are_duplicates():
                return
            if self.test_if_there_are_unrecognized_products():
                return
            self.preview_calculation()
        elif self.execute_button.text() == "Generar informe »":
            if self.test_if_code_missing():
                return
            self.generate_production_report()

    def test_if_there_are_no_products(self):
        no_line_items = not self.products_holder_section.children()
        contents = utils.get_line_items_contents(self.products_holder_section)
        empty_string = tests.test_if_empty_string_in_line_items_contents(contents)
        if no_line_items or empty_string:
            WarningBox(
                "Sin productos", "Completar o borrar campos\nvacíos antes de ejecutar."
            ).exec_()
            return True
        else:
            return False

    def test_if_there_are_duplicates(self):
        contents = utils.get_line_items_contents(self.products_holder_section)
        duplicates = tests.test_if_duplicated_first_value_in_line_items_contents(
            contents
        )
        if duplicates:
            WarningBox(
                "Componentes duplicados", "Borrar uno de los componentes duplicados."
            ).exec_()
            return True
        else:
            return False

    def test_if_there_are_unrecognized_products(self):
        contents = utils.get_line_items_contents(self.products_holder_section)
        products = contents[0::2]
        db = db_manager.DB_Manager()
        existing_recipes = db.get_all_recipes_as_display()
        db.close_connection()
        unrecognized_products = not set(products).issubset(existing_recipes)
        if unrecognized_products:
            WarningBox(
                "Componente extraño",
                "Componente no reconocido. Cargar el\ncomponente desde el autocompletado.",
            ).exec_()
            return True
        else:
            return False

    def test_if_code_missing(self):
        code_missing = self.code_field.text() == ""
        if code_missing:
            WarningBox("Código faltante", "Ingresar código\nantes de generar.").exec_()
            return True
        else:
            return False

    def preview_calculation(self):
        table = QtWidgets.QTableWidget()
        table.setFixedWidth(360)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        table.setFocusPolicy(QtCore.Qt.NoFocus)
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Componente", "Necesidad"])
        table.horizontalHeaderItem(0).setTextAlignment(QtCore.Qt.AlignHCenter)
        table.horizontalHeaderItem(1).setTextAlignment(QtCore.Qt.AlignHCenter)
        table.horizontalHeader().setDefaultSectionSize(70)
        table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        table_section = QtWidgets.QVBoxLayout()
        table_section.addWidget(table)

        self.layout.addLayout(table_section)

        self.add_button.setEnabled(False)

        utils.delete_layout(self.bottom_section)

        self.back_button = QtWidgets.QPushButton("« Volver")
        self.back_button.setShortcut("Alt+v")

        self.execute_button = QtWidgets.QPushButton("Calcular »")
        self.execute_button.setShortcut("Alt+c")

        self.bottom_section = QtWidgets.QHBoxLayout()
        self.bottom_section.addWidget(self.back_button)
        self.bottom_section.addWidget(self.execute_button)

        self.back_button.clicked.connect(self.close)
        self.add_button.clicked.connect(self.add_line_item)
        self.execute_button.clicked.connect(self.on_execute_clicked)

        self.layout.addLayout(self.bottom_section)

        self.entered_products_and_amounts = utils.get_line_items_contents(
            self.products_holder_section, as_dictionary=True
        )

        series_for_production = []

        db = db_manager.DB_Manager()

        for (
            product_to_be_manufactured,
            amount_to_be_manufactured,
        ) in self.entered_products_and_amounts.items():
            recipe_sql = utils.format_display_name_into_sql_name(
                product_to_be_manufactured
            )
            product_recipe = db.get_recipe_contents(recipe_sql)
            for comp_name, comp_amount in product_recipe.items():
                product_recipe[comp_name] = comp_amount * amount_to_be_manufactured

            series_for_production.append(product_recipe)

        db.close_connection()

        dd = defaultdict(list)

        for single_recipe in series_for_production:
            for component, amount in single_recipe.items():
                dd[component].append(amount)

        for key, value in dd.items():
            for _ in value:
                dd[key] = sum(value)

        self.needed_comps_and_amounts = dict(dd)

        table.setRowCount(len(self.needed_comps_and_amounts))

        utils.populate_table_column_with_list_of_strings(
            table=table, col_num=0, input_list=self.needed_comps_and_amounts.keys()
        )

        utils.populate_table_column_with_list_of_integers(
            table=table, col_num=1, input_list=self.needed_comps_and_amounts.values()
        )

        custom_height = table.rowCount() * 30 + 25
        if table.rowCount() <= 3:
            table.setMaximumHeight(custom_height)
            table.setFixedHeight(custom_height)
        elif table.rowCount() > 4:
            table.setMaximumHeight(205)
            table.setFixedHeight(205)

        self.execute_button.setText("Generar informe »")
        self.execute_button.setShortcut("Alt+g")

        self.code_field.setFocus()

        for l in self.products_holder_section.children():
            l.itemAt(0).widget().setDisabled(True)
            l.itemAt(1).widget().setDisabled(True)

        self.add_button.clicked.disconnect()

    def generate_production_report(self):
        code = self.code_field.text()
        path = os.getcwd() + "\\src\\main\\resources\\production_report_template.xlsx"
        wb = load_workbook(path)
        ws = wb["Informe"]
        for product_name, product_amount in self.entered_products_and_amounts.items():
            for cell in ws["B"][3:]:
                if cell.value is None:
                    cell.value = product_name
                    break

            for cell in ws["C"][3:]:
                if cell.value is None:
                    cell.value = utils.format_number_for_display(product_amount)
                    break

        db = db_manager.DB_Manager()
        df = db.make_needs_and_stocks_df(self.needed_comps_and_amounts)
        db.close_connection()
        assembler = code[0]
        fmt = utils.format_number_for_display
        for i in range(len(df)):
            for cell in ws["D"][3:]:
                if cell.value is None:
                    cell.value = df["comp"][i]
                    break

            for cell in ws["E"]:
                if cell.value is None:
                    cell.value = fmt(df["need"][i])
                    break

            for cell in ws["F"]:
                if cell.value is None:
                    if df["need"][i] <= df[assembler][i] + df["V"][i]:
                        cell.value = "Sí"
                    else:
                        diff = fmt(df["need"][i] - (df[assembler][i] + df["V"][i]))
                        cell.value = f"No, faltan {diff} unidades."
                    break

            for cell in ws["G"]:
                if cell.value is None:
                    cell.value = fmt(max(0, df["need"][i] - df[assembler][i]))
                    break

            for cell in ws["H"]:
                if cell.value is None:
                    cell.value = "0"
                    break

        ws["B2"] = code
        ws["F2"] = datetime.now().strftime("%d/%m/%Y")
        if code.startswith("K"):
            ws["H2"] = "Karina"
        elif code.startswith("B"):
            ws["H2"] = "Brid"
        elif code.startswith("T"):
            ws["H2"] = "Tercero"
        current_date_hyphens = datetime.now().strftime("%d-%m-%Y")
        filepath_and_name = (
            f"output\\INFORME DE PRODUCCIÓN {code} {current_date_hyphens}.xlsx"
        )
        wb.save(filepath_and_name)
        wb.close()
        os.startfile(
            f"output\\INFORME DE PRODUCCIÓN {code} {current_date_hyphens}.xlsx"
        )

        self.close()
