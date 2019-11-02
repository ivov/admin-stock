import os
from PyQt5 import QtWidgets, QtCore, QtGui
import openpyxl
from utils.styling import (
    generic_title_style,
    outgoing_dialog_table_left_title_style,
    outgoing_dialog_table_right_title_style,
    outgoing_applied_msgbox_title_1_style,
    outgoing_applied_msgbox_title_2_style,
    generic_messagebox_style,
)
from utils import db_manager
from utils import utils_collection as utils
from widgets.message_boxes import WarningBox


class OutgoingDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(OutgoingDialog, self).__init__(parent)
        self.setWindowFlags(
            QtCore.Qt.Dialog
            | QtCore.Qt.CustomizeWindowHint
            | QtCore.Qt.WindowCloseButtonHint
        )
        self.setFixedWidth(485)

        self.outgoing_comps = {}
        self.outgoing_needs = {}
        self.outgoing_comps_amts_needs = {}
        self.final_data_vald = {}
        self.final_data_assembler = {}
        self.code = ""
        self.assembler = ""

        screen_title = QtWidgets.QLabel("Egreso a armador")
        screen_title.setAlignment(QtCore.Qt.AlignCenter)
        screen_title.setStyleSheet(generic_title_style)

        self.filepicker_button = QtWidgets.QPushButton("Seleccionar informe...")
        self.filepicker_button.setShortcut("Alt+s")
        self.filepicker_button.setDefault(True)

        back_button = QtWidgets.QPushButton("« Volver")
        back_button.setShortcut("Alt+v")

        self.execute_button = QtWidgets.QPushButton("Ejecutar »")
        self.execute_button.setShortcut("Alt+e")
        self.execute_button.setEnabled(False)

        buttons_section = QtWidgets.QHBoxLayout()
        buttons_section.addWidget(back_button)
        buttons_section.addWidget(self.filepicker_button)
        buttons_section.addWidget(self.execute_button)

        self.layout = QtWidgets.QVBoxLayout()
        self.layout.addWidget(screen_title)
        self.layout.addLayout(buttons_section)

        self.setLayout(self.layout)

        back_button.clicked.connect(self.close)

        self.filepicker_button.clicked.connect(self.pick_file_and_build_table)
        self.execute_button.clicked.connect(self.insert_outgoing_comps_into_db)

    def pick_file_and_build_table(self):
        filepath, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Elegir informe", os.getcwd() + "\\output", "Excel (*.xlsx)"
        )
        if filepath == "":
            return
        file_name = filepath.split("/")[(-1)].replace(".xlsx", "")
        if not file_name.startswith("INFORME DE PRODUCCIÓN"):
            WarningBox(
                "Archivo equivocado",
                "El archivo seleccionado\nno es un informe de producción.",
            ).exec_()
            return

        self.filepicker_button.setEnabled(False)
        self.filepicker_button.setDefault(False)

        self.execute_button.setEnabled(True)
        self.execute_button.setDefault(True)

        report_name = file_name.replace("INFORME DE PRODUCCIÓN ", "").replace("-", "/")
        destinations = {"K": "Karina", "B": "Brid", "T": "Tercero"}
        destination_name = destinations[report_name[0]]

        filetitle = QtWidgets.QLabel("Informe seleccionado: " + report_name)
        filetitle.setStyleSheet(outgoing_dialog_table_left_title_style)

        destination = QtWidgets.QLabel("Destino: " + destination_name)
        destination.setStyleSheet(outgoing_dialog_table_right_title_style)

        table_title_section = QtWidgets.QHBoxLayout()
        table_title_section.addStretch()
        table_title_section.addWidget(filetitle)
        table_title_section.addWidget(destination)
        table_title_section.addStretch()

        self.layout.addLayout(table_title_section)

        table = QtWidgets.QTableWidget()
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        table.setColumnCount(3)
        table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        table.setHorizontalHeaderLabels(["Componente", "Necesidad", "Envío"])
        table.horizontalHeaderItem(0).setTextAlignment(QtCore.Qt.AlignHCenter)
        table.horizontalHeaderItem(1).setTextAlignment(QtCore.Qt.AlignHCenter)
        table.horizontalHeaderItem(2).setTextAlignment(QtCore.Qt.AlignHCenter)
        table.horizontalHeader().setDefaultSectionSize(110)
        table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["Informe"]
        self.code = ws["B"][1].value
        component_names = []

        for i in ws["D"][3:]:
            if i.value is not None:
                component_names.append(i.value)

        component_amounts = []
        for i in ws["H"][3:]:
            if i.value is not None:
                component_amounts.append(float(i.value))

        self.outgoing_comps = dict(zip(component_names, component_amounts))
        component_needs = []
        for i in ws["E"][3:]:
            if i.value is not None:
                component_needs.append(float(i.value))

        self.outgoing_needs = dict(zip(component_names, component_needs))
        table.setRowCount(len(self.outgoing_comps))
        utils.populate_table_column_with_list_of_strings(
            table=table, col_num=0, input_list=self.outgoing_comps.keys()
        )
        utils.populate_table_column_with_list_of_integers(
            table=table, col_num=1, input_list=self.outgoing_needs.values()
        )
        utils.populate_table_column_with_list_of_integers(
            table=table, col_num=2, input_list=self.outgoing_comps.values()
        )
        custom_height = table.rowCount() * 30 + 25
        if table.rowCount() <= 3:
            table.setMaximumHeight(custom_height)
            table.setFixedHeight(custom_height)
        elif table.rowCount() > 4:
            table.setMaximumHeight(205)
            table.setFixedHeight(205)

        self.layout.addWidget(table)

        self.center_dialog()

    def center_dialog(self):
        QtWidgets.QApplication.processEvents()
        screenGeometry = QtWidgets.QDesktopWidget().screenGeometry()
        x = screenGeometry.width() / 2 - self.width() / 2
        y = screenGeometry.height() / 2 - self.height() / 2
        self.move(x, y)

    def insert_outgoing_comps_into_db(self):
        self.outgoing_comps_amts_needs = dict(
            zip(
                self.outgoing_comps.keys(),
                list(zip(self.outgoing_comps.values(), self.outgoing_needs.values())),
            )
        )

        db = db_manager.DB_Manager()
        assemblers = {"K": "karina", "B": "brid", "T": "tercero"}

        self.assembler = assemblers[self.code[0]]

        for comp_display, amt_and_need in self.outgoing_comps_amts_needs.items():
            comp_sql = db.get_SQL_name_for_component(comp_display)
            comp_outgoing_amt = amt_and_need[0]
            comp_outgoing_need = amt_and_need[1]
            stock_vald_pre_appl = db.get_stock_at_valdenegro_for(comp_sql)
            stock_assembler_pre_appl = db.get_stock_at_assembler_for(
                comp_sql, self.assembler
            )
            stock_vald_post_appl = stock_vald_pre_appl - comp_outgoing_amt
            stock_assembler_post_appl = (
                stock_assembler_pre_appl + comp_outgoing_amt - comp_outgoing_need
            )
            self.final_data_vald[comp_display] = [
                stock_vald_pre_appl,
                stock_vald_post_appl,
            ]
            self.final_data_assembler[comp_display] = [
                stock_assembler_pre_appl,
                stock_assembler_post_appl,
            ]

            data = {
                "assembler": self.assembler,
                "comp_sql": comp_sql,
                "comp_outgoing_amt": comp_outgoing_amt,
                "comp_outgoing_need": comp_outgoing_need,
                "stock_vald_post_appl": stock_vald_post_appl,
                "stock_assembler_post_appl": stock_assembler_post_appl,
                "stock_karina": db.get_stock_at_assembler_for(comp_sql, "karina"),
                "stock_brid": db.get_stock_at_assembler_for(comp_sql, "brid"),
                "stock_tercero": db.get_stock_at_assembler_for(comp_sql, "tercero"),
            }

            db.apply_outgoing_to_valdenegro_and_assembler(data)
            settings = QtCore.QSettings("solutronic", "admin_stock")
            db.log_new_movement(
                movement="Egreso",
                destination=self.assembler.capitalize(),
                component=comp_display,
                amount=comp_outgoing_amt,
                user=settings.value("username"),
            )

        db.close_connection()

        self.close()

        OutgoingAppliedMessageBox(self).exec_()


class OutgoingAppliedMessageBox(QtWidgets.QMessageBox):
    def __init__(self, parent=None):
        super(OutgoingAppliedMessageBox, self).__init__(parent)

        self.setWindowTitle("Finalizado")
        self.setWindowFlags(
            QtCore.Qt.Dialog
            | QtCore.Qt.CustomizeWindowHint
            | QtCore.Qt.WindowTitleHint
            | QtCore.Qt.WindowCloseButtonHint
        )

        self.assembler = self.parent().assembler
        self.outgoing_comps = self.parent().outgoing_comps
        self.outgoing_needs = self.parent().outgoing_needs
        self.final_data_vald = self.parent().final_data_vald
        self.final_data_assembler = self.parent().final_data_assembler

        title_1 = QtWidgets.QLabel("Egreso aplicado a Depósito")
        title_1.setAlignment(QtCore.Qt.AlignCenter)
        title_1.setStyleSheet(outgoing_applied_msgbox_title_1_style)

        title_2 = QtWidgets.QLabel("Egreso aplicado a " + self.assembler)
        title_2.setAlignment(QtCore.Qt.AlignCenter)
        title_2.setStyleSheet(outgoing_applied_msgbox_title_2_style)

        table_1 = QtWidgets.QTableWidget()
        table_1.setFixedWidth(500)
        table_1.setFocusPolicy(QtCore.Qt.NoFocus)
        table_1.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table_1.verticalHeader().setVisible(False)
        table_1.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        table_1.setColumnCount(4)
        table_1.setHorizontalHeaderLabels(["Componente", "Inicial", "Egreso", "Final"])
        table_1.horizontalHeaderItem(0).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_1.horizontalHeaderItem(1).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_1.horizontalHeaderItem(2).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_1.horizontalHeaderItem(3).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_1.horizontalHeader().setDefaultSectionSize(70)
        table_1.horizontalHeader().setSectionResizeMode(
            0, QtWidgets.QHeaderView.Stretch
        )
        table_1.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        table_1.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        table_1.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        table_1.setRowCount(len(self.outgoing_comps))
        utils.populate_table_column_with_list_of_strings(
            table=table_1, col_num=0, input_list=self.outgoing_comps.keys()
        )

        components_pre_application_amts_vald = [
            i[0] for i in self.final_data_vald.values()
        ]

        utils.populate_table_column_with_list_of_integers(
            table=table_1, col_num=1, input_list=components_pre_application_amts_vald
        )

        utils.populate_table_column_with_list_of_integers(
            table=table_1,
            col_num=2,
            input_list=self.outgoing_comps.values(),
            with_minus=True,
        )

        for i in range(table_1.rowCount()):
            table_1.item(i, 2).setBackground(QtGui.QColor("red"))
            table_1.item(i, 2).setForeground(QtGui.QColor("white"))

        components_post_application_amts_vald = [
            i[1] for i in self.final_data_vald.values()
        ]

        utils.populate_table_column_with_list_of_integers(
            table=table_1, col_num=3, input_list=components_post_application_amts_vald
        )

        custom_height = table_1.rowCount() * 30 + 25

        table_1.setMaximumHeight(custom_height)
        table_1.setFixedHeight(custom_height)
        table_2 = QtWidgets.QTableWidget()
        table_2.setFocusPolicy(QtCore.Qt.NoFocus)
        table_2.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table_2.verticalHeader().setVisible(False)
        table_2.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        table_2.setColumnCount(5)
        table_2.setHorizontalHeaderLabels(
            ["Componente", "Inicial", "Recepción", "Necesidad", "Final"]
        )
        table_2.horizontalHeaderItem(0).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_2.horizontalHeaderItem(1).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_2.horizontalHeaderItem(2).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_2.horizontalHeaderItem(3).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_2.horizontalHeaderItem(4).setTextAlignment(QtCore.Qt.AlignHCenter)
        table_2.horizontalHeader().setDefaultSectionSize(70)
        table_2.horizontalHeader().setSectionResizeMode(
            0, QtWidgets.QHeaderView.Stretch
        )
        table_2.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        table_2.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        table_2.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        table_2.setRowCount(len(self.outgoing_comps))

        utils.populate_table_column_with_list_of_strings(
            table=table_2, col_num=0, input_list=self.outgoing_comps.keys()
        )

        components_pre_application_amts_assembler = [
            i[0] for i in self.final_data_assembler.values()
        ]

        utils.populate_table_column_with_list_of_integers(
            table=table_2,
            col_num=1,
            input_list=components_pre_application_amts_assembler,
        )

        utils.populate_table_column_with_list_of_integers(
            table=table_2,
            col_num=2,
            input_list=self.outgoing_comps.values(),
            with_plus=True,
        )

        for i in range(table_2.rowCount()):
            table_2.item(i, 2).setBackground(QtGui.QColor("#79c879"))

        utils.populate_table_column_with_list_of_integers(
            table=table_2,
            col_num=3,
            input_list=self.outgoing_needs.values(),
            with_minus=True,
        )

        for i in range(table_2.rowCount()):
            table_2.item(i, 3).setBackground(QtGui.QColor("red"))
            table_2.item(i, 3).setForeground(QtGui.QColor("white"))

        components_post_application_amts_assembler = [
            i[1] for i in self.final_data_assembler.values()
        ]

        utils.populate_table_column_with_list_of_integers(
            table=table_2,
            col_num=4,
            input_list=components_post_application_amts_assembler,
        )

        custom_height = table_2.rowCount() * 30 + 25

        table_2.setMaximumHeight(custom_height)
        table_2.setFixedHeight(custom_height)

        ok_button = QtWidgets.QPushButton("OK")
        ok_button.clicked.connect(self.close)

        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(title_1)
        layout.addWidget(table_1)
        layout.addWidget(title_2)
        layout.addWidget(table_2)
        layout.addWidget(ok_button)

        self.layout().addLayout(layout, 0, 0, 0, 0, QtCore.Qt.AlignCenter)
        self.setStyleSheet(generic_messagebox_style)

        admin_window = self.parent().parent().parent().parent()
        admin_window.statusbar.show_quick_message("Egreso aplicado")
        admin_window.start_screen.rebuild_main_section()
