from datetime import datetime
import os
from PyQt5 import QtWidgets, QtGui, QtCore
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from utils.styling import (
    comp_details_dialog_title_style,
    movements_dialog_table_view_style,
)
from utils import db_manager
from utils import utils_collection as utils


class ComponentDetailsDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(ComponentDetailsDialog, self).__init__(parent)

        self.selected_comp_display = self.parent().selected_comp_display
        self.selected_comp_SQL = self.parent().selected_comp_SQL

        self.setWindowFlags(
            QtCore.Qt.Dialog
            | QtCore.Qt.CustomizeWindowHint
            | QtCore.Qt.WindowCloseButtonHint
        )

        self.categories_excluded_from_table = [
            "current_stock_valdenegro",
            "current_stock_karina",
            "current_stock_brid",
            "current_stock_tercero",
        ]

        title = QtWidgets.QLabel(self.selected_comp_display)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet(comp_details_dialog_title_style)

        subtitles_layout = QtWidgets.QHBoxLayout()
        subtitle = QtWidgets.QLabel()
        subtitle.setStyleSheet(comp_details_dialog_title_style)

        db = db_manager.DB_Manager()
        self.comp_details = db.get_comp_details(self.selected_comp_SQL)
        db.close_connection()

        current_main = self.comp_details["current_stock_valdenegro"]
        current_first = self.comp_details["current_stock_karina"]
        current_second = self.comp_details["current_stock_brid"]
        current_third = self.comp_details["current_stock_tercero"]
        current_stocks = [current_main, current_first, current_second, current_third]
        subtitles = ["Stock Valdenegro", "Stock Karina", "Stock Brid"]
        stocks_and_subtitles = dict(zip(subtitles, current_stocks))
        subtitles_layout.addStretch()

        for subtitle, stock in stocks_and_subtitles.items():
            subtitle = QtWidgets.QLabel(subtitle + ": " + stock)
            subtitles_layout.addWidget(subtitle)
            subtitles_layout.addStretch()

        self.table = QtWidgets.QTableWidget()
        self.table.setFocusPolicy(QtCore.Qt.NoFocus)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.table.setRowCount(len(self.comp_details["fecha"]))
        self.table.setStyleSheet(movements_dialog_table_view_style)
        self.table.setColumnCount(16)
        self.fill_up_table()
        self.format_table()
        self.color_table()

        back_button = QtWidgets.QPushButton("« Volver")
        back_button.setShortcut("Alt+v")

        self.export_button = QtWidgets.QPushButton("Exportar detalles »")
        self.export_button.setShortcut("Alt+x")

        table_layout = QtWidgets.QVBoxLayout()
        table_layout.addWidget(self.table)

        bottom_section = QtWidgets.QHBoxLayout()
        bottom_section.addWidget(back_button)
        bottom_section.addWidget(self.export_button)

        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(title)
        layout.addLayout(subtitles_layout)
        layout.addLayout(table_layout)
        layout.addLayout(bottom_section)

        self.setLayout(layout)

        back_button.clicked.connect(self.close)
        self.export_button.clicked.connect(self.export_to_spreadsheet)

    def fill_up_table(self):
        self.table.setHorizontalHeaderLabels(
            [
                "Fecha",
                "Remito",
                "Proveedor",
                "Nota",
                "D: Ingreso",
                "D: Egreso",
                "D: Stock",
                "K: Ingreso",
                "K: Egreso",
                "K: Stock",
                "B: Ingreso",
                "B: Egreso",
                "B: Stock",
                "T: Ingreso",
                "T: Egreso",
                "T: Stock",
            ]
        )

        for count, category in enumerate(self.comp_details):
            if category not in self.categories_excluded_from_table:
                utils.populate_table_column_with_list_of_strings(
                    table=self.table,
                    col_num=count,
                    input_list=self.comp_details[category],
                )

    def format_table(self):
        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, 100)
        self.table.setColumnWidth(1, 60)
        self.table.setColumnWidth(2, 150)
        custom_width = 0

        for i in range(0, 16):
            custom_width += self.table.columnWidth(i)

        custom_width -= self.table.columnWidth(3)
        custom_width += 24
        self.setFixedWidth(custom_width)
        custom_height = self.table.rowCount() * 30 + 121
        if self.table.rowCount() <= 14:
            self.setMaximumHeight(custom_height)
            self.setFixedHeight(custom_height)
        elif self.table.rowCount() > 14:
            self.setMaximumHeight(578)
            self.setFixedHeight(578)

        last_item_index = len(self.comp_details) - 1
        self.table.ScrollHint(QtWidgets.QAbstractItemView.EnsureVisible)
        self.table.scrollToItem(
            self.table.item(last_item_index, 0),
            QtWidgets.QAbstractItemView.EnsureVisible,
        )
        self.table.hideColumn(3)
        self.table.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)

    def color_table(self):
        column_and_color = {
            "tan": QtGui.QColor("#f5deb3"),
            "pink": QtGui.QColor("#ffb6c1"),
            "light blue": QtGui.QColor("#00b7ff"),
            "brown": QtGui.QColor("#a6a62b"),
            "yellow": QtGui.QColor("#ffd700"),
        }

        for col in range(4):
            for row in range(self.table.rowCount()):
                self.table.item(row, col).setBackground(column_and_color["tan"])

        for col in range(4, 7):
            for row in range(self.table.rowCount()):
                self.table.item(row, col).setBackground(column_and_color["pink"])

        for col in range(7, 10):
            for row in range(self.table.rowCount()):
                self.table.item(row, col).setBackground(column_and_color["light blue"])

        for col in range(10, 13):
            for row in range(self.table.rowCount()):
                self.table.item(row, col).setBackground(column_and_color["brown"])

        for col in range(13, 16):
            for row in range(self.table.rowCount()):
                self.table.item(row, col).setBackground(column_and_color["yellow"])

    def export_to_spreadsheet(self):
        self.export_button.setText("Esperar...")
        path = os.getcwd() + "\\src\\main\\resources\\component_detail_template.xlsx"
        wb = load_workbook(path)
        ws = wb["Nombre"]
        ws.title = self.selected_comp_display
        ws["D"][0].value = self.selected_comp_display
        spreadsheet_columns = [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "I",
            "J",
            "K",
            "M",
            "N",
            "O",
            "Q",
            "R",
            "S",
        ]

        columns_and_datalist = dict(
            zip(spreadsheet_columns, self.comp_details.values())
        )

        for column, datalist in columns_and_datalist.items():
            for datapoint in datalist:
                for cell in ws[column][3:]:
                    if cell.value is None:
                        cell.value = datapoint
                        cell.alignment = Alignment(horizontal="right")
                        break

        current_date = datetime.now().strftime("%d-%m-%Y")
        filepath_and_name = (
            f"output\\Detalle de {self.selected_comp_display} {current_date}.xlsx"
        )
        wb.save(filepath_and_name)
        wb.close()
        os.startfile(filepath_and_name)
        self.export_button.setText("Exportar detalles »")
